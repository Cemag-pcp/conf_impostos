"""
Automação de Verificação de NF — Innovaro ERP

Etapa 1: login -> abrir o relatório "TI-Relatório de Baixas Recursos com
Chave NFe e Impostos" (Venda > Consultas) -> preencher período (Emissão
Início/Fim) -> clicar "Visualizar" -> coletar a tabela renderizada e
formatar como planilha (Excel).

Etapa 2: abrir o relatório "99003 Download de XML Manifestados (C)" (Fiscal
e Regulamentação > Consultas > Auxiliares Fiscais > Manifestação (C)) ->
preencher período (Data Inicial/Final) -> clicar "Executar" -> baixar o
.zip com os XMLs das NFes manifestadas no período.

As telas de filtros dos dois relatórios usam Shadow DOM fechado (ver
BOAS_PRATICAS_INNOVARO_PLAYWRIGHT.md, seção 25) — interage via CDP
(cdp_helpers.py). Já a tabela/link de RESULTADO de cada relatório (depois de
"Visualizar"/"Executar") é DOM normal — dá pra ler/clicar direto com
Playwright.

Etapa 3: descompacta o .zip de XMLs (etapa 2), lê cada XML de NFe (formato
padrão `nfeProc` da SEFAZ) e monta uma planilha com um resumo por nota e o
detalhamento por item — pura leitura de arquivo, não depende do navegador.
"""

import os
import time
import logging
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, Page, TimeoutError as PlaywrightTimeout

from cdp_helpers import find_node, focus_node, click_text

# Diretório do próprio script — NUNCA usar caminho relativo "puro" (relativo
# ao diretório de trabalho) neste projeto. Como a automação roda sozinha,
# agendada de madrugada, o diretório de trabalho no momento da execução pode
# não ser a pasta do projeto (ex.: Agendador de Tarefas do Windows às vezes
# inicia com "Iniciar em" vazio) — sem isso, .env, downloads/ e output/
# seriam resolvidos no lugar errado ou simplesmente não seriam encontrados.
# Resolver tudo a partir de BASE_DIR garante que funciona em qualquer
# computador e a partir de qualquer diretório de onde o script for chamado.
BASE_DIR = Path(__file__).resolve().parent

load_dotenv(BASE_DIR / ".env")


def _caminho_config(valor: str, base: Path = BASE_DIR) -> Path:
    """Resolve um caminho de configuração relativo a BASE_DIR (a menos que já seja absoluto)."""
    p = Path(valor)
    return p if p.is_absolute() else base / p


# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------

ERP_URL          = os.getenv("INNOVARO_URL", "http://192.168.3.140/sistema")
ERP_USER         = os.getenv("INNOVARO_USERNAME", "")
ERP_PASS         = os.getenv("INNOVARO_PASSWORD", "")
SLOW_MO          = int(os.getenv("INNOVARO_SLOW_MO_MS", "250"))
VIEWPORT_WIDTH   = int(os.getenv("INNOVARO_VIEWPORT_WIDTH", "1366"))
VIEWPORT_HEIGHT  = int(os.getenv("INNOVARO_VIEWPORT_HEIGHT", "800"))
DOWNLOAD_DIR     = _caminho_config(os.getenv("DOWNLOAD_DIR", "downloads"))

NOME_RELATORIO = "TI-Relatório de Baixas Recursos com Chave NFe e Impostos"
CAMINHO_RELATORIO = "Venda > Consultas"

NOME_RELATORIO_XML = "99003 Download de XML Manifestados (C)"
CAMINHO_RELATORIO_XML = "Fiscal e Regulamentação > Consultas > Auxiliares Fiscais > Manifestação (C)"

GOOGLE_SHEETS_ID = os.getenv("GOOGLE_SHEETS_ID", "")
GOOGLE_SHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

OUTPUT_DIR = BASE_DIR / "output" / "screenshots"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(BASE_DIR / "output" / "automacao.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Utilitários
# ---------------------------------------------------------------------------

def capturar_screenshot(page: Page, nome: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho = OUTPUT_DIR / f"{ts}_{nome}.png"
    page.screenshot(path=str(caminho), full_page=True)
    return caminho


def aguardar_erp_pronto(page: Page, wait_ms: int = 500, timeout_ms: int = 60000):
    """Aguarda a barra de progresso global do Innovaro sair do estado ocupado."""
    page.wait_for_load_state("domcontentloaded")
    try:
        page.wait_for_load_state("networkidle", timeout=5000)
    except Exception:
        pass

    barra = page.locator('[role="progressbar"]').first
    inicio = time.time()
    while (time.time() - inicio) * 1000 < timeout_ms:
        try:
            if barra.count() == 0:
                break
            classes = barra.get_attribute("class") or ""
            if "mdc-linear-progress--closed" in classes:
                break
        except Exception:
            break
        page.wait_for_timeout(300)
    else:
        raise TimeoutError("ERP não terminou de processar dentro do tempo esperado.")

    if wait_ms > 0:
        page.wait_for_timeout(wait_ms)


def aguardar_acao_completar(page: Page, timeout_ocupado_ms: int = 5000, timeout_total_ms: int = 120000, wait_ms: int = 500):
    """
    Espera "inteligente" para ações que disparam processamento no servidor
    (ex.: clicar em "Visualizar" num relatório, que pode demorar um tempo
    variável para gerar). Duas fases:

      1. Espera a barra de progresso global FICAR OCUPADA — confirma que a
         ação realmente começou a processar no servidor. Sem essa fase, uma
         checagem de "está ocioso?" logo após o clique pode ver a barra
         ainda no estado ocioso ANTERIOR (a ação nem começou a processar) e
         seguir em frente cedo demais, achando que já terminou.
      2. Espera a barra voltar a ficar OCIOSA — o tempo de carregamento é
         dinâmico (varia com o volume de dados do relatório), por isso não
         se usa um sleep fixo aqui.
    """
    barra = page.locator('[role="progressbar"]').first

    inicio = time.time()
    while (time.time() - inicio) * 1000 < timeout_ocupado_ms:
        try:
            if barra.count() == 0:
                break
            classes = barra.get_attribute("class") or ""
            if "mdc-linear-progress--closed" not in classes:
                break  # ficou ocupado — ação começou a processar
        except Exception:
            break
        page.wait_for_timeout(100)

    aguardar_erp_pronto(page, wait_ms=wait_ms, timeout_ms=timeout_total_ms)


# ---------------------------------------------------------------------------
# Login
# ---------------------------------------------------------------------------

def login(page: Page):
    log.info("Fazendo login...")
    page.goto(ERP_URL, wait_until="domcontentloaded")
    page.wait_for_timeout(1000)

    try:
        page.get_by_role("textbox", name="Usuário").fill(ERP_USER)
        page.get_by_role("textbox", name="Senha").fill(ERP_PASS)
        page.get_by_role("button", name="Entrar", exact=True).click()
    except PlaywrightTimeout:
        page.locator("#username").fill(ERP_USER)
        page.locator("#password").fill(ERP_PASS)
        page.locator("#submit-login").click()

    page.get_by_role("button", name="Menu", exact=True).wait_for(timeout=30000)
    log.info("Login realizado com sucesso.")


# ---------------------------------------------------------------------------
# Pesquisa do sistema (dialog global de busca)
# ---------------------------------------------------------------------------

def abrir_pesquisa_sistema(page: Page):
    campo = page.get_by_role("textbox", name="Pesquisa do sistema")
    if campo.is_visible():
        return
    page.get_by_role("button", name="Pesquisar", exact=True).click()
    campo.wait_for(state="visible", timeout=10000)


def pesquisar_no_sistema(page: Page, termo: str):
    """IMPORTANTE: usar keyboard.type(), .fill() não dispara a busca (ver BOAS_PRATICAS seção 23)."""
    abrir_pesquisa_sistema(page)
    campo = page.get_by_role("textbox", name="Pesquisa do sistema")
    campo.click()
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    page.wait_for_timeout(200)
    page.keyboard.type(termo, delay=40)
    page.wait_for_timeout(1500)


def abrir_relatorio(page: Page, nome_exato: str, timeout_ms: int = 15000):
    """Abre um relatório pelo nome exato (listado na pesquisa do sistema)."""
    log.info(f"Abrindo relatório '{nome_exato}'...")
    pesquisar_no_sistema(page, nome_exato)
    item = page.get_by_role("listitem").filter(has=page.get_by_text(nome_exato, exact=True))
    item.first.wait_for(state="visible", timeout=timeout_ms)
    item.first.click()
    aguardar_erp_pronto(page, wait_ms=1000)
    log.info("Relatório aberto.")


# ---------------------------------------------------------------------------
# Formulário de filtros (Shadow DOM fechado — via CDP)
# ---------------------------------------------------------------------------

def _esperar_no(page: Page, cdp, predicate, timeout_ms: int = 8000):
    """
    Espera até `predicate(tag, attrs)` encontrar algum nó via CDP (o Shadow
    DOM da tela pode levar um instante para ser montado/atualizado logo após
    abrir uma tela nova). Retorna o nó ou None se esgotar o timeout.
    """
    inicio = time.time()
    while (time.time() - inicio) * 1000 < timeout_ms:
        achado = find_node(cdp, predicate)
        if achado is not None:
            return achado
        page.wait_for_timeout(300)
    return None


def preencher_data(page: Page, cdp, nome_campo: str, data_ddmmyyyy: str):
    """
    Preenche um campo de data do formulário de filtros do relatório.
    `data_ddmmyyyy` no formato DDMMAAAA (só dígitos) — o campo formata sozinho.
    """
    node = _esperar_no(page, cdp, lambda tag, a: tag == "INPUT" and a.get("name") == nome_campo)
    if node is None:
        raise RuntimeError(f"Campo de filtro '{nome_campo}' não encontrado.")
    focus_node(cdp, node)
    page.keyboard.press("Control+A")
    page.keyboard.type(data_ddmmyyyy, delay=50)
    page.keyboard.press("Tab")
    page.wait_for_timeout(500)


def visualizar_relatorio(page: Page, cdp, emissao_inicio: str, emissao_fim: str):
    """
    Preenche o período (Emissão Início/Fim) e clica em "Visualizar".
    Datas no formato DDMMAAAA (ex.: "25082026").
    """
    log.info(f"Preenchendo período: {emissao_inicio} até {emissao_fim}...")
    preencher_data(page, cdp, "dsvfilter_BAI_EMISSAO_START", emissao_inicio)
    preencher_data(page, cdp, "dsvfilter_BAI_EMISSAO_END", emissao_fim)

    log.info("Clicando em 'Visualizar'...")
    click_text(cdp, "Visualizar")
    aguardar_acao_completar(page, wait_ms=1500, timeout_total_ms=120000)


def executar_download_xml(page: Page, cdp, data_inicial: str, data_final: str):
    """
    Preenche o período (Data Inicial/Final) e clica em "Executar" no
    relatório "99003 Download de XML Manifestados (C)". Datas no formato
    DDMMAAAA. Diferente do relatório de baixas, os campos aqui não seguem o
    padrão dsvfilter_* — são só `dataInicial` / `dataFinal`.
    """
    log.info(f"Preenchendo período: {data_inicial} até {data_final}...")
    preencher_data(page, cdp, "dataInicial", data_inicial)
    preencher_data(page, cdp, "dataFinal", data_final)

    log.info("Clicando em 'Executar'...")
    click_text(cdp, "Executar")
    aguardar_acao_completar(page, wait_ms=1500, timeout_total_ms=120000)


def baixar_zip_xml(page: Page, timeout_ms: int = 30000) -> Path | None:
    """
    Depois de executar_download_xml, a tela de resultado mostra um ícone/link
    real (DOM normal, não Shadow DOM — <a href="http://monitor.nfemonitor.com.br/dfe?zip=...">
    target="_blank") para baixar um .zip com os XMLs das NFes manifestadas no
    período. Clica nele, captura o download e salva em DOWNLOAD_DIR.

    Duas formas de "sem resultados" aqui, tratadas as duas como None (mesmo
    padrão do relatório de Baixas) em vez de travar o script até dar timeout:

      1. Igual ao relatório de Baixas, quando não há nenhuma NFe manifestada
         no período o ERP mostra "A consulta não retornou resultados." em
         vez do link.
      2. Mais sutil: às vezes o ÍCONE de download aparece mesmo sem nenhum
         XML pronto pro período — é uma peça fixa da tela, não depende dos
         dados. Nesse caso o servidor externo (monitor.nfemonitor.com.br)
         nunca entrega o arquivo (resposta vazia, sem nunca disparar o
         evento de download) — confirmado ao vivo aguardando >90s sem o
         .zip nunca ficar pronto. Só dá pra saber tentando o download e
         esperando um tempo limitado pelo evento de fato.

    O clique pode disparar o download na própria aba OU abrir uma aba nova
    (o link é target="_blank") dependendo de como o servidor responde — por
    isso o evento "download" é escutado no nível do `context` (pega os dois
    casos), não em `page.expect_download()` (só pegaria o primeiro caso).
    """
    frame = obter_frame_relatorio(page)

    inicio = time.time()
    while (time.time() - inicio) * 1000 < timeout_ms:
        try:
            texto = frame.evaluate("() => document.body.innerText || ''")
        except Exception:
            texto = ""
        if TEXTO_SEM_RESULTADOS in texto.lower():
            log.info("  → Relatório de XMLs não retornou resultados para o período informado.")
            return None
        if frame.locator("a").count() > 0:
            break
        page.wait_for_timeout(300)
    else:
        raise TimeoutError("Link de download do ZIP não apareceu dentro do tempo esperado.")

    link = frame.locator("a").first
    link.wait_for(state="visible", timeout=timeout_ms)

    context = page.context
    download_holder = {}

    def _capturar(download):
        download_holder["download"] = download

    context.on("download", _capturar)
    try:
        paginas_antes = set(context.pages)
        link.click()
        inicio_download = time.time()
        while (time.time() - inicio_download) * 1000 < timeout_ms and "download" not in download_holder:
            page.wait_for_timeout(500)
    finally:
        context.remove_listener("download", _capturar)

    if "download" not in download_holder:
        # Fecha qualquer aba nova em branco que o clique tenha aberto, pra
        # não deixar lixo aberto no navegador.
        for pagina in set(context.pages) - paginas_antes:
            try:
                pagina.close()
            except Exception:
                pass
        # WARNING (não info) de propósito: diferente do "sem resultados" do
        # relatório de Baixas (mensagem explícita do ERP, caso confirmado),
        # aqui não dá pra saber com certeza se realmente não há XML ou se o
        # serviço externo (monitor.nfemonitor.com.br) só está indisponível
        # no momento — confirmado ao vivo (26/08/2026) esse serviço devolver
        # resposta vazia por mais de 1 minuto mesmo num período que tinha
        # XML de verdade minutos antes. Fica marcado como aviso pra alguém
        # notar no log e conferir manualmente se for um dia com movimento.
        log.warning(
            "  → Link de download não gerou nenhum arquivo dentro do tempo esperado. "
            "Pode ser que não haja XML pronto para o período, OU que o serviço externo "
            "de download (monitor.nfemonitor.com.br) esteja indisponível no momento — "
            "conferir manualmente se o período tinha movimento. Seguindo com 'Resumo NFe'/'Itens' vazias."
        )
        return None

    download = download_holder["download"]
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    caminho = DOWNLOAD_DIR / download.suggested_filename
    download.save_as(str(caminho))
    log.info(f"  → ZIP de XMLs salvo em: {caminho}")
    return caminho


# ---------------------------------------------------------------------------
# Coleta da tabela de resultado (DOM normal, dentro do iframe da aba)
# ---------------------------------------------------------------------------

def obter_frame_relatorio(page: Page):
    iframe = page.locator("iframe").last
    iframe.wait_for(timeout=20000)
    frame = iframe.element_handle().content_frame()
    if frame is None:
        raise RuntimeError("Não foi possível acessar o iframe do relatório.")
    return frame


# Trecho (sem acento, pra não depender de encoding) da mensagem que o ERP
# mostra em vez de tabela/link quando o período filtrado não tem nenhum
# registro. Como o script agora roda de madrugada sempre com hoje como único
# dia do período, esse é um resultado ROTINEIRO (não um erro) — precisa ser
# tratado como "0 linhas", nunca deixar a coleta esperar até dar timeout.
TEXTO_SEM_RESULTADOS = "retornou resultados"

# Cabeçalho do relatório de Baixas, fixo, usado como fallback de esquema
# quando o relatório não retorna nenhuma linha (não dá pra descobrir o
# cabeçalho lendo a tabela, porque não existe tabela nesse caso). Mantém as
# abas seguintes (Cruzamento etc.) funcionando normalmente mesmo com 0 linhas
# em "Baixas", já que elas dependem de saber o nome/posição das colunas.
COLUNAS_BAIXAS_PADRAO = [
    "Emissão", "Pessoa.Código", "Recurso", "Recurso.Classe.Nome", "Quantidade",
    "Total", "Número", "Ch Acesso DFe", "Pes Brut Vols", "Pes Líq Vols",
    "Qde Volumes", "PIS", "PIS ST", "Acréscimo Item", "Acréscimo", "IPI",
    "BC ICMS", "ICMS", "BC ICMS ST", "BC PIS", "COFINS", "BC COFINS",
    "COFINS ST", "BC ISS", "ISS", "BC INSS", "INSS", "BC CBS", "CBS", "BC IBS",
]


def coletar_tabela_relatorio(page: Page, timeout_ms: int = 30000) -> pd.DataFrame:
    """
    Lê a tabela renderizada pelo relatório (após visualizar_relatorio) e
    retorna como DataFrame.

    A tabela HTML mistura três tipos de linha:
      - linhas de cabeçalho/metadados do relatório (empresa, data de geração,
        filtros aplicados) — no topo, com poucas células;
      - a linha de CABEÇALHO DE COLUNA de verdade (primeira célula == "Emissão");
      - linhas de DETALHE (mesma quantidade de células do cabeçalho);
      - linhas de SUBTOTAL por grupo (ex.: "▾Total", "▾ALMOÇO" — só 3
        células: rótulo do grupo, soma de Quantidade, soma de Total).

    Só as linhas de DETALHE (mesma largura do cabeçalho) entram no resultado.
    Quando o período não tem nenhum registro, o ERP mostra a mensagem "A
    consulta não retornou resultados." em vez da tabela — nesse caso retorna
    um DataFrame vazio (0 linhas) com o esquema de colunas padrão, em vez de
    esperar até dar timeout.
    """
    # Busca o frame DE NOVO a cada tentativa — clicar em "Visualizar" pode
    # recarregar/trocar o iframe da aba, e uma referência pega cedo demais
    # (antes da troca) fica "detached" e derruba o frame.evaluate().
    inicio = time.time()
    linhas = []
    while (time.time() - inicio) * 1000 < timeout_ms:
        try:
            frame = obter_frame_relatorio(page)
            texto = frame.evaluate("() => document.body.innerText || ''")
            if TEXTO_SEM_RESULTADOS in texto.lower():
                log.info("  → Relatório não retornou resultados para o período informado.")
                return pd.DataFrame(columns=COLUNAS_BAIXAS_PADRAO)
            linhas = frame.evaluate("""() => {
                return [...document.querySelectorAll('tr')].map(
                    tr => [...tr.children].map(td => td.textContent.trim())
                );
            }""")
        except Exception as exc:
            log.warning(f"  → Tentativa de ler a tabela falhou ({exc}) — tentando de novo...")
            page.wait_for_timeout(500)
            continue
        # Considera "pronto" quando já existe uma linha cujo primeiro valor é
        # exatamente "Emissão" (o cabeçalho de coluna real do relatório).
        if any(linha and linha[0] == "Emissão" for linha in linhas):
            break
        page.wait_for_timeout(500)
    else:
        raise TimeoutError("Tabela do relatório não carregou dentro do tempo esperado.")

    idx_cabecalho = next(i for i, linha in enumerate(linhas) if linha and linha[0] == "Emissão")
    cabecalho = linhas[idx_cabecalho]
    n_colunas = len(cabecalho)

    dados = [
        linha for linha in linhas[idx_cabecalho + 1:]
        if len(linha) == n_colunas
    ]

    log.info(f"  → Tabela coletada: {len(dados)} linha(s) de detalhe, {n_colunas} colunas.")
    return pd.DataFrame(dados, columns=cabecalho)


# ---------------------------------------------------------------------------
# Formatação / saída
# ---------------------------------------------------------------------------

def converter_numero_br(valor: str):
    """Converte número no formato brasileiro ('1.234,56') para float. Vazio -> None."""
    valor = (valor or "").strip()
    if not valor:
        return None
    try:
        return float(valor.replace(".", "").replace(",", "."))
    except ValueError:
        return valor  # não era número (ex.: texto), mantém como está


COLUNAS_NUMERICAS = [
    "Quantidade", "Total", "Pes Brut Vols", "Pes Líq Vols", "Qde Volumes",
    "PIS", "PIS ST", "Acréscimo Item", "Acréscimo", "IPI", "BC ICMS", "ICMS",
    "BC ICMS ST", "BC PIS", "COFINS", "BC COFINS", "COFINS ST", "BC ISS",
    "ISS", "BC INSS", "INSS", "BC CBS", "CBS", "BC IBS",
]


def formatar_tabela(df: pd.DataFrame) -> pd.DataFrame:
    """Converte colunas numéricas (formato BR) e datas para tipos nativos."""
    df = df.copy()
    for col in COLUNAS_NUMERICAS:
        if col in df.columns:
            df[col] = df[col].apply(converter_numero_br)
    if "Emissão" in df.columns:
        df["Emissão"] = pd.to_datetime(df["Emissão"], format="%d/%m/%Y", errors="coerce")
    return df


# ---------------------------------------------------------------------------
# Processamento dos XMLs de NFe (etapa 3 — não usa o navegador)
# ---------------------------------------------------------------------------

NFE_NS = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

# Esquemas de coluna padrão de "Resumo NFe"/"Itens" (mesmas chaves usadas em
# parsear_nfe_xml), usados como fallback quando não há nenhum XML pra
# processar (nenhuma NFe manifestada no período) — mantém "Cruzamento" e as
# demais abas funcionando normalmente mesmo com 0 linhas.
COLUNAS_RESUMO_PADRAO = [
    "Arquivo", "Chave NFe", "Número", "Série", "Data Emissão", "Natureza Operação",
    "Emitente CNPJ", "Emitente Nome", "Destinatário CNPJ", "Destinatário Nome",
    "Valor Produtos", "Valor ICMS", "Valor IPI", "Valor PIS", "Valor COFINS",
    "Valor Frete", "Valor Desconto", "Valor Total NF", "Situação", "Protocolo",
    "Data Autorização",
]
COLUNAS_ITENS_PADRAO = [
    "Chave NFe", "Item", "Código Produto", "Descrição", "NCM", "CFOP",
    "Unidade", "Quantidade", "Valor Unitário", "Valor Total Item",
]


def descompactar_zip(caminho_zip: Path) -> Path:
    """Descompacta o .zip de XMLs numa subpasta com o mesmo nome (sem extensão)."""
    destino = caminho_zip.with_suffix("")
    destino.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(caminho_zip) as z:
        z.extractall(destino)
    n = len(list(destino.glob("*.xml")))
    log.info(f"  → {n} XML(s) extraído(s) em: {destino}")
    return destino


def _texto(elem, caminho: str, default=""):
    """Busca o texto de um sub-elemento (caminho com namespace 'nfe:'), ou default se não achar/vazio."""
    if elem is None:
        return default
    node = elem.find(caminho, NFE_NS)
    return node.text if node is not None and node.text is not None else default


def _numero(elem, caminho: str, default=None):
    valor = _texto(elem, caminho, None)
    if valor is None:
        return default
    try:
        return float(valor)
    except ValueError:
        return default


def parsear_nfe_xml(caminho_xml: Path):
    """
    Lê um XML de NFe (formato padrão `nfeProc` da SEFAZ) e retorna
    (resumo: dict, itens: list[dict]) — um resumo por nota e uma linha por
    item/produto dentro dela.
    """
    root = ET.parse(caminho_xml).getroot()

    inf_nfe = root.find(".//nfe:infNFe", NFE_NS)
    chave = (inf_nfe.get("Id") or "").replace("NFe", "") if inf_nfe is not None else ""

    ide = root.find(".//nfe:ide", NFE_NS)
    emit = root.find(".//nfe:emit", NFE_NS)
    dest = root.find(".//nfe:dest", NFE_NS)
    total = root.find(".//nfe:total/nfe:ICMSTot", NFE_NS)
    prot = root.find(".//nfe:protNFe/nfe:infProt", NFE_NS)

    resumo = {
        "Arquivo": caminho_xml.name,
        "Chave NFe": chave,
        "Número": _texto(ide, "nfe:nNF"),
        "Série": _texto(ide, "nfe:serie"),
        "Data Emissão": _texto(ide, "nfe:dhEmi"),
        "Natureza Operação": _texto(ide, "nfe:natOp"),
        "Emitente CNPJ": _texto(emit, "nfe:CNPJ"),
        "Emitente Nome": _texto(emit, "nfe:xNome"),
        "Destinatário CNPJ": _texto(dest, "nfe:CNPJ"),
        "Destinatário Nome": _texto(dest, "nfe:xNome"),
        "Valor Produtos": _numero(total, "nfe:vProd"),
        "Valor ICMS": _numero(total, "nfe:vICMS"),
        "Valor IPI": _numero(total, "nfe:vIPI"),
        "Valor PIS": _numero(total, "nfe:vPIS"),
        "Valor COFINS": _numero(total, "nfe:vCOFINS"),
        "Valor Frete": _numero(total, "nfe:vFrete"),
        "Valor Desconto": _numero(total, "nfe:vDesc"),
        "Valor Total NF": _numero(total, "nfe:vNF"),
        "Situação": _texto(prot, "nfe:xMotivo"),
        "Protocolo": _texto(prot, "nfe:nProt"),
        "Data Autorização": _texto(prot, "nfe:dhRecbto"),
    }

    itens = []
    for det in root.findall(".//nfe:det", NFE_NS):
        prod = det.find("nfe:prod", NFE_NS)
        itens.append({
            "Chave NFe": chave,
            "Item": det.get("nItem"),
            "Código Produto": _texto(prod, "nfe:cProd"),
            "Descrição": _texto(prod, "nfe:xProd"),
            "NCM": _texto(prod, "nfe:NCM"),
            "CFOP": _texto(prod, "nfe:CFOP"),
            "Unidade": _texto(prod, "nfe:uCom"),
            "Quantidade": _numero(prod, "nfe:qCom"),
            "Valor Unitário": _numero(prod, "nfe:vUnCom"),
            "Valor Total Item": _numero(prod, "nfe:vProd"),
        })

    return resumo, itens


def coletar_dados_xmls(pasta_xmls: Path):
    """Percorre todos os .xml de uma pasta e monta os DataFrames de resumo (1 linha/NFe) e itens (1 linha/produto)."""
    arquivos = sorted(pasta_xmls.glob("*.xml"))
    log.info(f"  → Processando {len(arquivos)} arquivo(s) XML...")

    resumos, todos_itens = [], []
    for arq in arquivos:
        try:
            resumo, itens = parsear_nfe_xml(arq)
            resumos.append(resumo)
            todos_itens.extend(itens)
        except Exception as exc:
            log.warning(f"  → Falha ao processar '{arq.name}': {exc}")

    df_resumo = pd.DataFrame(resumos)
    for col in ("Data Emissão", "Data Autorização"):
        if col in df_resumo.columns:
            df_resumo[col] = pd.to_datetime(df_resumo[col], errors="coerce", utc=True).dt.tz_localize(None)

    df_itens = pd.DataFrame(todos_itens)
    return df_resumo, df_itens


def _forcar_texto(df: pd.DataFrame, coluna: str) -> pd.DataFrame:
    """
    Garante que uma coluna de "chave" (ex.: chave de acesso de NFe, 44
    dígitos) seja gravada como TEXTO no Excel, nunca como número.

    Achado crítico: se a coluna passar em algum momento por
    `pd.read_excel()` (ex.: reaproveitando um arquivo já salvo em vez de
    reler o ERP), o pandas pode inferir esses valores como `int` do Python
    (que tem precisão arbitrária, então parece correto). Só na hora de
    salvar de novo é que o Excel converte pra número de ponto flutuante
    (float64, ~15-17 dígitos de precisão) — uma chave de 44 dígitos perde
    os dígitos finais e vira outro número (silenciosamente, sem erro). O
    pipeline normal (raspagem do ERP -> DataFrame -> Excel) nunca passa por
    esse round-trip, mas essa blindagem evita o problema em qualquer reuso
    futuro do código (ex.: rodar só a Etapa 3 a partir de um Excel salvo).
    """
    if coluna in df.columns:
        df = df.copy()
        df[coluna] = df[coluna].astype(str)
    return df


def salvar_planilha_final(df_baixas: pd.DataFrame, df_resumo: pd.DataFrame, df_itens: pd.DataFrame) -> Path:
    """Salva as três tabelas num único Excel, cada uma na sua aba."""
    df_baixas = _forcar_texto(df_baixas, "Ch Acesso DFe")
    df_resumo = _forcar_texto(df_resumo, "Chave NFe")
    df_itens = _forcar_texto(df_itens, "Chave NFe")

    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho = DOWNLOAD_DIR / f"verificacao_nf_{ts}.xlsx"
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df_baixas.to_excel(writer, sheet_name="Baixas", index=False)
        df_resumo.to_excel(writer, sheet_name="Resumo NFe", index=False)
        df_itens.to_excel(writer, sheet_name="Itens", index=False)
    log.info(f"  → Planilha final salva em: {caminho} (abas: Baixas, Resumo NFe, Itens)")
    return caminho


def fluxo_processar_xmls(caminho_zip: Path | None, df_baixas: pd.DataFrame) -> Path:
    """
    Descompacta o .zip de XMLs, extrai os dados de cada NFe e salva a
    planilha final com quatro abas: Baixas (Etapa 1), Resumo NFe e Itens
    (extraídos dos XMLs, Etapa 2), e Cruzamento (fórmulas comparando Baixas
    x Resumo NFe por imposto).

    `caminho_zip` pode ser None (nenhuma NFe manifestada no período) — nesse
    caso "Resumo NFe"/"Itens" ficam vazias (0 linhas), mas com o esquema de
    colunas padrão, pra planilha final sair completa mesmo assim.
    """
    if caminho_zip is None:
        log.info("  → Nenhum XML manifestado no período — 'Resumo NFe' e 'Itens' ficarão vazias.")
        df_resumo = pd.DataFrame(columns=COLUNAS_RESUMO_PADRAO)
        df_itens = pd.DataFrame(columns=COLUNAS_ITENS_PADRAO)
    else:
        pasta = descompactar_zip(caminho_zip)
        df_resumo, df_itens = coletar_dados_xmls(pasta)
        log.info(f"  → {len(df_resumo)} NFe(s), {len(df_itens)} item(ns) no total.")
    caminho = salvar_planilha_final(df_baixas, df_resumo, df_itens)
    montar_aba_cruzamento(caminho, df_baixas, df_resumo)

    try:
        enviar_para_google_sheets(df_baixas, df_resumo, df_itens)
    except Exception:
        # A planilha local (o entregável principal) já foi salva nesse ponto
        # — uma falha no envio ao Sheets (rede, permissão, aba renomeada)
        # não deve derrubar a execução agendada de madrugada.
        log.exception("  → Falha ao enviar dados para o Google Sheets (planilha local já salva normalmente).")

    return caminho


# ---------------------------------------------------------------------------
# Google Sheets — acrescenta as linhas de cada execução (histórico
# acumulado; roda 1x/dia, cada execução só cobre o dia de hoje)
# ---------------------------------------------------------------------------

def _credenciais_google_sheets() -> dict:
    """Monta o dict de credenciais de service account a partir das variáveis GOOGLE_* do .env."""
    return {
        "type": os.getenv("GOOGLE_TYPE"),
        "project_id": os.getenv("GOOGLE_PROJECT_ID"),
        "private_key_id": os.getenv("GOOGLE_PRIVATE_KEY_ID"),
        "private_key": os.getenv("GOOGLE_PRIVATE_KEY"),
        "client_email": os.getenv("GOOGLE_CLIENT_EMAIL"),
        "client_id": os.getenv("GOOGLE_CLIENT_ID"),
        "auth_uri": os.getenv("GOOGLE_AUTH_URI"),
        "token_uri": os.getenv("GOOGLE_TOKEN_URI"),
        "auth_provider_x509_cert_url": os.getenv("GOOGLE_AUTH_PROVIDER_CERT_URL"),
        "client_x509_cert_url": os.getenv("GOOGLE_CLIENT_CERT_URL"),
    }


def _preparar_linhas_sheets(df: pd.DataFrame) -> list:
    """
    Converte um DataFrame pra lista de listas serializável pela API do Sheets:
    datas viram texto (dd/mm/aaaa) e NaN/NaT viram string vazia (o Sheets
    rejeita float('nan') — não é um valor JSON válido).
    """
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime("%d/%m/%Y")
    df = df.astype(object).where(pd.notnull(df), "")
    return df.values.tolist()


def enviar_para_google_sheets(df_baixas: pd.DataFrame, df_resumo: pd.DataFrame, df_itens: pd.DataFrame):
    """
    Acrescenta (não sobrescreve) as linhas desta execução nas abas "Baixas",
    "Resumo NFe" e "Itens" da planilha Google configurada em GOOGLE_SHEETS_ID
    — histórico acumulado dia a dia, já que o script roda 1x/dia e cada
    execução só cobre o período de hoje. Escreve o cabeçalho primeiro se a
    aba estiver totalmente vazia (primeira execução).
    """
    if not GOOGLE_SHEETS_ID:
        log.info("  → GOOGLE_SHEETS_ID não configurado — pulando envio ao Google Sheets.")
        return

    import gspread
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_info(_credenciais_google_sheets(), scopes=GOOGLE_SHEETS_SCOPES)
    cliente = gspread.authorize(creds)
    planilha = cliente.open_by_key(GOOGLE_SHEETS_ID)

    abas = {"Baixas": df_baixas, "Resumo NFe": df_resumo, "Itens": df_itens}
    for nome_aba, df in abas.items():
        aba = planilha.worksheet(nome_aba)
        if not aba.get_all_values():
            aba.append_row(df.columns.tolist(), value_input_option="USER_ENTERED")
        linhas = _preparar_linhas_sheets(df)
        if linhas:
            aba.append_rows(linhas, value_input_option="USER_ENTERED")
        log.info(f"  → Aba '{nome_aba}' do Google Sheets: {len(linhas)} linha(s) acrescentada(s).")


# ---------------------------------------------------------------------------
# Aba "Cruzamento" — fórmulas do Excel, Baixas x Resumo NFe por imposto
# ---------------------------------------------------------------------------

# Impostos presentes nas DUAS abas (a "Baixas" tem outros — BC ISS, INSS, CBS,
# IBS etc. — mas sem contrapartida no XML da NFe, então ficam fora do cruzamento).
IMPOSTOS_CRUZAMENTO = ["ICMS", "IPI", "PIS", "COFINS"]


def _letra_coluna(df: pd.DataFrame, nome_coluna: str) -> str:
    return get_column_letter(df.columns.get_loc(nome_coluna) + 1)


def montar_aba_cruzamento(caminho_xlsx: Path, df_baixas: pd.DataFrame, df_resumo: pd.DataFrame):
    """
    Adiciona a aba "Cruzamento" com FÓRMULAS do Excel (não valores fixos)
    comparando, por NFe (chave de acesso), a soma dos impostos lançados na
    aba "Baixas" (SUMIF por "Ch Acesso DFe") contra o valor informado no XML
    da nota (aba "Resumo NFe"). A aba "Itens" fica de fora — é só consulta,
    não entra no cruzamento.

    Planilha "automática": uma linha do Cruzamento por linha de "Resumo
    NFe" (mesmo número de linha — Cruzamento!linha N sempre lê 'Resumo
    NFe'!linha N), e todas as fórmulas usam REFERÊNCIA DE COLUNA INTEIRA
    (ex.: Baixas!$H:$H, não um intervalo fixo tipo $H$2:$H$6). Isso é o que
    permite simplesmente ARRASTAR a última linha pra baixo quando entrarem
    mais NFes em "Resumo NFe" no futuro — não precisa rodar o script de novo
    pra "religar" as fórmulas.

    Existência da NFe em "Baixas" é checada por COUNTIF antes de comparar:
    Status = "SEM BAIXA" (não achou a chave em Baixas — não é divergência de
    valor, só ainda não foi lançada), "OK" (achou e os valores batem) ou
    "DIVERGENTE" (achou mas algum imposto ou o total não bate).
    """
    col_baixas_chave = _letra_coluna(df_baixas, "Ch Acesso DFe")
    col_baixas_total = _letra_coluna(df_baixas, "Total")
    cols_baixas_impostos = {imp: _letra_coluna(df_baixas, imp) for imp in IMPOSTOS_CRUZAMENTO}

    col_resumo_chave = _letra_coluna(df_resumo, "Chave NFe")
    col_resumo_numero = _letra_coluna(df_resumo, "Número")
    col_resumo_emitente = _letra_coluna(df_resumo, "Emitente Nome")
    col_resumo_total = _letra_coluna(df_resumo, "Valor Total NF")
    cols_resumo_impostos = {imp: _letra_coluna(df_resumo, f"Valor {imp}") for imp in IMPOSTOS_CRUZAMENTO}

    ref_baixas_chave = f"'Baixas'!${col_baixas_chave}:${col_baixas_chave}"
    ref_baixas_total = f"'Baixas'!${col_baixas_total}:${col_baixas_total}"
    ref_baixas_impostos = {imp: f"'Baixas'!${c}:${c}" for imp, c in cols_baixas_impostos.items()}

    wb = load_workbook(caminho_xlsx)
    ws = wb.create_sheet("Cruzamento")

    cabecalho = ["Chave NFe", "Número", "Emitente", "Tem Baixa?"]
    for imp in IMPOSTOS_CRUZAMENTO:
        cabecalho += [f"{imp} (Baixas)", f"{imp} (NFe)", f"Diferença {imp}"]
    cabecalho += ["Total (Baixas)", "Total (NFe)", "Diferença Total", "Status"]
    ws.append(cabecalho)

    for i in range(len(df_resumo)):
        linha = i + 2  # linha 1 = cabeçalho; Cruzamento!linha == 'Resumo NFe'!linha sempre
        row = [
            f"='Resumo NFe'!{col_resumo_chave}{linha}",
            f"='Resumo NFe'!{col_resumo_numero}{linha}",
            f"='Resumo NFe'!{col_resumo_emitente}{linha}",
            f"=IF(COUNTIF({ref_baixas_chave},A{linha})>0,\"SIM\",\"NÃO\")",
        ]
        colunas_diferenca = []

        for imp in IMPOSTOS_CRUZAMENTO:
            f_baixas = f"=SUMIF({ref_baixas_chave},A{linha},{ref_baixas_impostos[imp]})"
            f_nfe = f"='Resumo NFe'!{cols_resumo_impostos[imp]}{linha}"
            letra_b = get_column_letter(len(row) + 1)
            letra_n = get_column_letter(len(row) + 2)
            f_dif = f"=ROUND({letra_b}{linha}-{letra_n}{linha},2)"
            row += [f_baixas, f_nfe, f_dif]
            colunas_diferenca.append(get_column_letter(len(row)))

        letra_total_b = get_column_letter(len(row) + 1)
        row.append(f"=SUMIF({ref_baixas_chave},A{linha},{ref_baixas_total})")
        letra_total_n = get_column_letter(len(row) + 1)
        row.append(f"='Resumo NFe'!{col_resumo_total}{linha}")
        letra_total_dif = get_column_letter(len(row) + 1)
        row.append(f"=ROUND({letra_total_b}{linha}-{letra_total_n}{linha},2)")
        colunas_diferenca.append(letra_total_dif)

        soma_abs = "+".join(f"ABS({c}{linha})" for c in colunas_diferenca)
        col_tem_baixa = "D"
        row.append(f'=IF({col_tem_baixa}{linha}="NÃO","SEM BAIXA",IF({soma_abs}<0.01,"OK","DIVERGENTE"))')

        ws.append(row)

    wb.save(caminho_xlsx)
    log.info(
        f"  → Aba 'Cruzamento' adicionada em: {caminho_xlsx} "
        f"({len(df_resumo)} linha(s), 1 por NFe de 'Resumo NFe' — fórmulas com coluna inteira, "
        f"arrastar pra baixo funciona pra linhas novas; impostos: {', '.join(IMPOSTOS_CRUZAMENTO)})."
    )


# ---------------------------------------------------------------------------
# Fluxo completo
# ---------------------------------------------------------------------------

def fluxo_baixas_recursos_nfe(page: Page, cdp, emissao_inicio: str, emissao_fim: str) -> pd.DataFrame:
    """
    Abre o relatório "TI-Relatório de Baixas Recursos com Chave NFe e
    Impostos", filtra pelo período informado, coleta e formata a tabela.
    """
    abrir_relatorio(page, NOME_RELATORIO)
    visualizar_relatorio(page, cdp, emissao_inicio, emissao_fim)
    df = coletar_tabela_relatorio(page)
    return formatar_tabela(df)


def fluxo_download_xml_manifestados(page: Page, cdp, data_inicial: str, data_final: str) -> Path | None:
    """
    Abre o relatório "99003 Download de XML Manifestados (C)", filtra pelo
    período informado, executa e baixa o .zip com os XMLs das NFes. Retorna
    None se não houver nenhuma NFe manifestada no período.
    """
    abrir_relatorio(page, NOME_RELATORIO_XML)
    executar_download_xml(page, cdp, data_inicial, data_final)
    return baixar_zip_xml(page)


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------

def main():
    # Roda desacompanhado (agendado de madrugada) — sempre usa a data de hoje
    # como início e fim do período, sem depender de argumento nenhum.
    hoje = datetime.now().strftime("%d%m%Y")
    emissao_inicio = emissao_fim = hoje

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log.info("Iniciando automação de Verificação de NF...")

    with sync_playwright() as p:
        browser = p.chromium.launch(channel="chrome", headless=False, slow_mo=SLOW_MO)
        context = browser.new_context(viewport={"width": VIEWPORT_WIDTH, "height": VIEWPORT_HEIGHT})
        page = context.new_page()
        cdp = context.new_cdp_session(page)
        cdp.send("DOM.enable")

        try:
            login(page)

            log.info("=== Etapa 1: TI-Relatório de Baixas Recursos com Chave NFe e Impostos ===")
            df_baixas = fluxo_baixas_recursos_nfe(page, cdp, emissao_inicio, emissao_fim)
            log.info(f"Etapa 1 concluída — {len(df_baixas)} linha(s) coletada(s).")

            log.info("=== Etapa 2: 99003 Download de XML Manifestados (C) ===")
            caminho_zip = fluxo_download_xml_manifestados(page, cdp, emissao_inicio, emissao_fim)
            if caminho_zip is None:
                log.info("Etapa 2 concluída — nenhuma NFe manifestada no período.")
            else:
                log.info(f"Etapa 2 concluída — XMLs baixados em {caminho_zip}.")
        except Exception:
            capturar_screenshot(page, "erro_geral")
            raise
        finally:
            browser.close()

    log.info("=== Etapa 3: Processar XMLs baixados e montar planilha final ===")
    caminho_planilha = fluxo_processar_xmls(caminho_zip, df_baixas)
    log.info(f"Etapa 3 concluída — planilha em {caminho_planilha}.")


if __name__ == "__main__":
    main()
